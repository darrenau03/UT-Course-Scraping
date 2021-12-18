from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

'''
PATH: path of chrome web driver
COURSE_WEBSITE: website for course schedule for specified semester
USERNAME: UT EID
PASSWORD: password
'''

PATH = "C:\Program Files (x86)\chromedriver.exe"
COURSE_WEBSITE = "https://utdirect.utexas.edu/apps/registrar/course_schedule/20222/"
USERNAME = ""
PASSWORD = ""


options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)
driver = webdriver.Chrome(options=options, executable_path=PATH)

# search the website
driver.get(COURSE_WEBSITE)

# sign in
search = driver.find_element_by_id("username")
search.send_keys(USERNAME)
search = driver.find_element_by_id("password")
search.send_keys(PASSWORD)
driver.find_element_by_name("_eventId_proceed").click()

wait = WebDriverWait(driver, 100)
wait.until(EC.presence_of_element_located((By.ID, "main_title")))
fields_driver = Select(driver.find_element_by_id("fos_cn"))

dict = {}
for a in fields_driver.options:
    dict[a.text.split(' - ')[0]] = a.text

PATH_COURSES = "courses.txt"

# read list of colleges
f = open(PATH_COURSES)
str = f.read()
course_names = str.split("\n")
print(course_names)
f.close()

wb = Workbook()
wb.remove(wb['Sheet'])

for i in course_names:
    sheet = wb.create_sheet(i)
    sheet.title = i
    driver.get(COURSE_WEBSITE)

    wait = WebDriverWait(driver, 100)
    wait.until(EC.presence_of_element_located((By.ID, "main_title")))
    fields_driver = Select(driver.find_element_by_id("fos_cn"))

    number_driver = driver.find_element_by_id("course_number")

    # select the class from dropdown
    field = i.rsplit(" ", 1)[0]
    fields_driver.select_by_visible_text(dict[field])
    number_driver.send_keys(i.rsplit(" ", 1)[1])
    find_courses = driver.find_elements_by_class_name("submit_button")[4]
    find_courses.click()

    # list of rows
    course = driver.find_elements_by_tag_name("tr")
    print("num courses found", len(course))
    c = 1
    for k in course:
        # info in each individual row
        try:
            info = k.find_elements_by_tag_name("td")

            # accessing the info
            unique = info[0].find_element_by_tag_name("a")
            sheet.cell(row=c, column=1).value = unique.text
            print(unique.text)

            days = info[1].find_elements_by_tag_name("span")
            s = ""
            for j in days:
                s += " " + j.text
            sheet.cell(row=c, column=4).value = s

            hours = info[2].find_elements_by_tag_name("span")
            s = ""
            for j in hours:
                s += " " + j.text
            sheet.cell(row=c, column=5).value = s

            # room = info[3].find_elements_by_tag_name("span")
            # for j in room:
            #     print(j.text)
            status = info[6].text
            sheet.cell(row=c, column=3).value = status
            print(status)

            try:
                instructor = info[5].find_element_by_tag_name("span")
                sheet.cell(row=c, column=2).value = instructor.text.lower()
                print(instructor.text.lower())
            except Exception as e:
                print(e)

            c += 1

        except Exception as e:

            print(e)

    teachers = []
    for i in range(1, c):
        teacher = sheet.cell(row=i, column=2).value
        if teacher not in teachers and teacher != None:
            teachers.append(teacher)

    print(teachers)
    d = 0
    for i in teachers:
        sheet.cell(row=c + 2 + d, column=1).value = i
        d += 1
driver.close()
wb.save("output.xlsx")
